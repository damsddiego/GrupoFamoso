# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from time import gmtime, strftime
from datetime import datetime
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from json import dumps
import ast
import json
from io import BytesIO


class TaxReportWizardSales(models.TransientModel):
    _name = "tax.report.wizard.sales"
    _description = "Sales Tax Report"

    start_date = fields.Date(string="Start Date", required=True)
    end_date = fields.Date(string="End Date", default=datetime.today(), required=True)
    report_for = fields.Selection([
        ('partners', 'Partners'),
        ('productCategory', 'Product Category'),
        ('products', 'Products'),
        ('sales_team', 'Sales Team'),
        ('sales_person', 'Salesperson'),
        ('taxes', 'Taxes')
    ], string='Report for')
    company_ids = fields.Many2many(
        comodel_name='res.company',
        relation='res_company_sales_rel',
        column1='wizard_id',
        column2='company_id',
        string="Company",
        required=True,
        default=lambda self: self.env.company,
        domain="[('id', 'in', context.get('allowed_company_ids'))]"
    )
    detailed_report = fields.Boolean(string='Detailed Report')
    report_format = fields.Selection([
        ('excel', 'Excel'),
        ('pdf', 'PDF'),
    ], string='Report format', default='excel', required=True)

    partner_ids = fields.Many2many(
        comodel_name='res.partner',
        relation='res_partner_sales_rel',
        column1='wizard_id',
        column2='partner_id',
        string='Partners'
    )
    product_category_ids = fields.Many2many(comodel_name='product.category', string='Product Categories')
    product_ids = fields.Many2many(comodel_name='product.product', string='Products')
    sales_team_ids = fields.Many2many(comodel_name='crm.team', string='Sales Teams')
    salesperson_ids = fields.Many2many(comodel_name='res.users', string='Sales Persons')
    tax_ids = fields.Many2many(
        comodel_name='account.tax',
        relation='account_tax_sales_rel',
        column1='wizard_id',
        column2='tax_id',
        string='Taxes'
    )

    tax_ids_domain = fields.Text(string="Sales domain", default=[])
    sales_person_ids_domain = fields.Text(string="SalesPersons domain", default=[])
    generic_domain = fields.Text(string="common domain", default=[])
    partner_ids_domain = fields.Text(string="Partners domain", default=[])

    @api.onchange('company_ids')
    def _onchange_companyIds_taxType(self):
        domain = []
        domain.append(('company_id.id', 'in', self.company_ids.ids))
        domain.append(('type_tax_use', '=', 'sale'))
        self.tax_ids_domain = domain
        self.tax_ids = [(5, 0, 0)]

    @api.onchange('company_ids')
    def _onchange_companyIds(self):
        domain = ['|', ('company_id', '=', False), ('company_id.id', 'in', self.company_ids.ids)]
        self.generic_domain = domain
        self.partner_ids = [(5, 0, 0)]
        self.product_ids = [(5, 0, 0)]
        self.sales_team_ids = [(5, 0, 0)]
        self.sales_person_ids_domain = [('company_ids', 'in', self.company_ids.ids)]
        self.partner_ids_domain = [('customer_rank', '>', 0)] + domain
        self.salesperson_ids = [(5, 0, 0)]

    def _filter_inv_on_tax(self, tax, invoice):
        line = []
        for inv in invoice:
            in_out_refund = inv.move_type in ['out_refund']
            taxes = inv.tax_totals
            if taxes['subtotals'][0]['tax_groups']:
                for tax_group in taxes['subtotals'][0]['tax_groups']:
                    if tax_group.get('group_name') == tax:
                        # Get all amount fields from invoice with safe access
                        amount_discount_electronic_invoice = getattr(inv, 'amount_discount_electronic_invoice', 0.0) or 0.0
                        amount_iva_returned = getattr(inv, 'amount_iva_returned', 0.0) or 0.0
                        amount_subtotal_without_iva_returned = getattr(inv, 'amount_subtotal_without_iva_returned', 0.0) or 0.0
                        amount_tax_signed = getattr(inv, 'amount_tax_signed', 0.0) or 0.0
                        amount_total_signed = getattr(inv, 'amount_total_signed', 0.0) or 0.0
                        state_tributacion = getattr(inv, 'state_tributacion', '') or ''

                        # Apply refund logic to amounts
                        discount_adjusted = amount_discount_electronic_invoice if not in_out_refund else -abs(amount_discount_electronic_invoice)
                        total_signed_adjusted = amount_total_signed if not in_out_refund else -abs(amount_total_signed)

                        # Calculate Sub Total (Total Signed + Discount Electronic Invoice)
                        sub_total = total_signed_adjusted + discount_adjusted

                        val = {
                            'name': inv.partner_id.name,
                            'ref': inv.name,
                            'inv_date': inv.invoice_date.strftime("%d/%m/%y"),
                            'sub_total': sub_total,
                            'amount_discount_electronic_invoice': discount_adjusted,
                            'amount_iva_returned': amount_iva_returned if not in_out_refund else -abs(amount_iva_returned),
                            'amount_subtotal_without_iva_returned': amount_subtotal_without_iva_returned if not in_out_refund else -abs(amount_subtotal_without_iva_returned),
                            'amount_tax_signed': amount_tax_signed if not in_out_refund else -abs(amount_tax_signed),
                            'amount_total_signed': total_signed_adjusted,
                            'state_tributacion': state_tributacion,
                        }
                        line.append(val)
        return line

    def _get_base_domain(self):
        """Returns base domain for sales invoices"""
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ('move_type', 'in', ['out_invoice', 'out_refund']),
        ]
        return domain

    def get_report_datas_partners(self):
        domain = self._get_base_domain()
        if self.partner_ids:
            domain.append(('partner_id.id', 'in', self.partner_ids.ids))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("There are not any invoices, Please review your selection."))

        partner_list = invoice_lines.mapped('partner_id').ids
        partners = self.env['res.partner'].sudo().search([('id', 'in', partner_list)])

        data_dics = []
        for p in partners:
            invoice = invoice_lines.filtered(lambda a: a.partner_id.id == p.id)
            bal_total_untax = 0
            total_tax_collected = 0
            balance_total_tax = 0
            tax_groups = []
            partner_dic = {}
            tex_gr_names = []  # Resetear para cada partner para evitar duplicados

            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund']
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not in_out_refund else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            balance_total_tax += tax_group.get('tax_amount') if not in_out_refund else -abs(tax_group.get('tax_amount'))
                            total_tax_collected += tax_group.get('tax_amount') if not in_out_refund else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name, invoice),
                })

            partner_dic['name'] = p.name if p.name else 'Name Not Available'
            partner_dic['bal_total_untax'] = bal_total_untax
            partner_dic['total_tax_collected'] = total_tax_collected
            partner_dic['balance_total_tax'] = balance_total_tax
            partner_dic['tax_groups'] = tax_groups
            data_dics.append(partner_dic)

        return data_dics

    def get_report_datas_taxes(self):
        domain = self._get_base_domain()
        if self.tax_ids:
            domain.append(('invoice_line_ids.tax_ids', 'in', self.tax_ids.ids))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("There are not any invoices, Please review your selection."))

        tax_lst = invoice_lines.mapped('invoice_line_ids').mapped('tax_ids').ids
        tax_ids = self.env['account.tax'].sudo().search([('id', 'in', tax_lst)])

        data_dics = []

        for tax in tax_ids:
            invoice = invoice_lines.filtered(lambda a: tax in a.invoice_line_ids.tax_ids)
            bal_total_untax = 0
            total_tax_collected = 0
            balance_total_tax = 0
            tax_groups = []
            tax_dic = {}
            tex_gr_names = []  # Resetear para cada tax para evitar duplicados

            for inv in invoice:
                in_out_refund = inv.move_type in ['out_refund']
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        bal_total_untax += inv.amount_untaxed if not in_out_refund else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            balance_total_tax += tax_group.get('tax_amount') if not in_out_refund else -abs(tax_group.get('tax_amount'))
                            total_tax_collected += tax_group.get('tax_amount') if not in_out_refund else -abs(tax_group.get('tax_amount'))

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name, invoice),
                })

            tax_dic['name'] = tax.name if tax.name else 'Name Not Available'
            tax_dic['bal_total_untax'] = bal_total_untax
            tax_dic['total_tax_collected'] = total_tax_collected
            tax_dic['balance_total_tax'] = balance_total_tax
            tax_dic['tax_groups'] = tax_groups
            data_dics.append(tax_dic)

        return data_dics

    def _write_cell(self, worksheet, row, col, value, style_name):
        """Helper to write cell with style"""
        cell = worksheet.cell(row=row, column=col)
        cell.value = value

        if 'header' in style_name:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        elif 'bold' in style_name:
            cell.font = Font(bold=True)

        if 'center' in style_name:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif 'right' in style_name:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif 'left' in style_name:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    def print_xlsx_report(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sales Tax Report'

        row = 1
        col = 1

        # Title
        self._write_cell(worksheet, row, col, 'Sales Tax Report', 'header_center_bold')
        row += 2

        # Date range
        self._write_cell(worksheet, row, col, f'From: {self.start_date} To: {self.end_date}', 'left_bold')
        row += 2

        if self.report_for == 'taxes':
            # Headers
            self._write_cell(worksheet, row, col, _('Tax'), 'header_center_bold')
            self._write_cell(worksheet, row, col+1, 'Total Untaxed Amount', 'header_right') if not self.detailed_report else None
            self._write_cell(worksheet, row, col+2, 'Total IVA Cobrado', 'header_right') if not self.detailed_report else None
            self._write_cell(worksheet, row, col+3, 'Balance', 'header_right') if not self.detailed_report else None
            row += 1

            data = self.get_report_datas_taxes()

            # Totales generales para toda la tabla
            grand_totals = {
                'sub_total': 0,
                'amount_discount_electronic_invoice': 0,
                'amount_iva_returned': 0,
                'amount_subtotal_without_iva_returned': 0,
                'amount_tax_signed': 0,
                'amount_total_signed': 0,
            }

            for rec in data:
                self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, rec.get('total_tax_collected'), 'right_bold') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                row += 1

                for tex_line in rec.get('tax_groups'):
                    if self.detailed_report:
                        self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                        row += 1

                        # Headers específicos para VENTAS
                        self._write_cell(worksheet, row, col+2, _("Cliente"), 'header_center_bold')
                        self._write_cell(worksheet, row, col+3, _("Factura"), 'header_center_bold')
                        self._write_cell(worksheet, row, col+4, _("Fecha Factura"), 'header_center_bold')
                        self._write_cell(worksheet, row, col+5, _("Sub Total"), 'header_right')
                        self._write_cell(worksheet, row, col+6, _("Descuento Factura Electrónica"), 'header_right')
                        self._write_cell(worksheet, row, col+7, _("Subtotal Sin IVA Devuelto"), 'header_right')
                        self._write_cell(worksheet, row, col+8, _("IVA Cobrado"), 'header_right')
                        self._write_cell(worksheet, row, col+9, _("IVA Devuelto"), 'header_right')
                        self._write_cell(worksheet, row, col+10, _("Total Firmado"), 'header_right')
                        self._write_cell(worksheet, row, col+11, _("Estado Tributación"), 'header_center_bold')
                        row += 1

                        # Initialize totals
                        totals = {
                            'sub_total': 0,
                            'amount_discount_electronic_invoice': 0,
                            'amount_iva_returned': 0,
                            'amount_subtotal_without_iva_returned': 0,
                            'amount_tax_signed': 0,
                            'amount_total_signed': 0,
                        }

                        for line in tex_line.get('lines'):
                            self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                            self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                            self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                            self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                            self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                            self._write_cell(worksheet, row, col+7, line.get('amount_subtotal_without_iva_returned', 0), 'right')
                            self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                            self._write_cell(worksheet, row, col+9, line.get('amount_iva_returned', 0), 'right')
                            self._write_cell(worksheet, row, col+10, line.get('amount_total_signed', 0), 'right')
                            self._write_cell(worksheet, row, col+11, line.get('state_tributacion', ''), 'left')

                            # Update totals del grupo
                            for key in totals.keys():
                                totals[key] += line.get(key, 0)

                            # Update totales generales
                            for key in grand_totals.keys():
                                grand_totals[key] += line.get(key, 0)

                            row += 1

                        # Write totals row
                        self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                        self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                        self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                        self._write_cell(worksheet, row, col+7, totals['amount_subtotal_without_iva_returned'], 'right_bold')
                        self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                        self._write_cell(worksheet, row, col+9, totals['amount_iva_returned'], 'right_bold')
                        self._write_cell(worksheet, row, col+10, totals['amount_total_signed'], 'right_bold')
                        row += 1
                    else:
                        self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                        total_base = 0
                        total_tax = 0
                        for line in tex_line.get('lines'):
                            total_base += line.get('base_amount', 0) or 0
                            total_tax += line.get('tax_amount', 0) or 0
                        self._write_cell(worksheet, row, col+2, total_base, 'right')
                        self._write_cell(worksheet, row, col+3, total_tax, 'right')
                        row += 1
                row += 1

            # Escribir fila de TOTALES GENERALES al final (solo en reporte detallado)
            if self.detailed_report:
                row += 1
                # Aplicar color de fondo amarillo para resaltar la fila de totales
                self._write_cell(worksheet, row, col+2, 'TOTAL GENERAL', 'header_center_bold')
                self._write_cell(worksheet, row, col+5, grand_totals['sub_total'], 'header_right')
                self._write_cell(worksheet, row, col+6, grand_totals['amount_discount_electronic_invoice'], 'header_right')
                self._write_cell(worksheet, row, col+7, grand_totals['amount_subtotal_without_iva_returned'], 'header_right')
                self._write_cell(worksheet, row, col+8, grand_totals['amount_tax_signed'], 'header_right')
                self._write_cell(worksheet, row, col+9, grand_totals['amount_iva_returned'], 'header_right')
                self._write_cell(worksheet, row, col+10, grand_totals['amount_total_signed'], 'header_right')

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)

        export_id = self.env['account.excel.tax.report'].create({
            'excel_file': base64.b64encode(fp.getvalue()),
            'file_name': 'sales_tax_report.xlsx'
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'account.excel.tax.report',
            'view_mode': 'form',
            'res_id': export_id.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

    def print_report(self):
        self.ensure_one()
        if self.report_format == 'excel':
            return self.print_xlsx_report()
        elif self.report_format == 'pdf':
            # TODO: Implement PDF report
            raise models.UserError(_("PDF report not implemented yet"))
