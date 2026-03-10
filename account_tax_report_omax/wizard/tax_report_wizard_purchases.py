# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from time import gmtime, strftime
from datetime import datetime
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from json import dumps
import ast
import json
from io import BytesIO


class TaxReportWizardPurchases(models.TransientModel):
    _name = "tax.report.wizard.purchases"
    _description = "Purchases Tax Report"

    start_date = fields.Date(string="Start Date", required=True)
    end_date = fields.Date(string="End Date", default=datetime.today(), required=True)
    report_for = fields.Selection([
        ('partners', 'Partners'),
        ('productCategory', 'Product Category'),
        ('products', 'Products'),
        ('taxes', 'Taxes')
    ], string='Report for')
    company_ids = fields.Many2many(
        comodel_name='res.company',
        relation='res_company_purchases_rel',
        column1='wizard_id',
        column2='company_id',
        string="Company",
        required=True,
        default=lambda self: self.env.company,
        domain="[('id', 'in', context.get('allowed_company_ids'))]"
    )
    detailed_report = fields.Boolean(string='Detailed Report')
    report_format = fields.Selection([
        ('excel', 'Excel'),
        ('pdf', 'PDF'),
    ], string='Report format', default='excel', required=True)

    partner_ids = fields.Many2many(
        comodel_name='res.partner',
        relation='res_partner_purchases_rel',
        column1='wizard_id',
        column2='partner_id',
        string='Suppliers'
    )
    product_category_ids = fields.Many2many(comodel_name='product.category', string='Product Categories')
    product_ids = fields.Many2many(comodel_name='product.product', string='Products')
    tax_ids = fields.Many2many(
        comodel_name='account.tax',
        relation='account_tax_purchases_rel',
        column1='wizard_id',
        column2='tax_id',
        string='Taxes'
    )

    tax_ids_domain = fields.Text(string="Purchases domain", default=[])
    generic_domain = fields.Text(string="common domain", default=[])
    partner_ids_domain = fields.Text(string="Partners domain", default=[])

    @api.onchange('company_ids')
    def _onchange_companyIds_taxType(self):
        domain = []
        domain.append(('company_id.id', 'in', self.company_ids.ids))
        domain.append(('type_tax_use', '=', 'purchase'))
        self.tax_ids_domain = domain
        self.tax_ids = [(5, 0, 0)]

    @api.onchange('company_ids')
    def _onchange_companyIds(self):
        domain = ['|', ('company_id', '=', False), ('company_id.id', 'in', self.company_ids.ids)]
        self.generic_domain = domain
        self.partner_ids = [(5, 0, 0)]
        self.product_ids = [(5, 0, 0)]
        self.partner_ids_domain = [('supplier_rank', '>', 0)] + domain

    def _get_invoice_tax_breakdown(self, inv):
        """
        Obtiene el desglose de impuestos por tasa para una factura
        Retorna un dict con las tasas y montos base y tax
        """
        in_out_refund = inv.move_type in ['in_refund']
        multiplier = 1 if not in_out_refund else -1

        # Diccionario para almacenar montos por tasa de impuesto
        tax_breakdown = {}

        # Recorrer las líneas de la factura para obtener impuestos
        for line in inv.invoice_line_ids:
            for tax in line.tax_ids:
                tax_rate = int(tax.amount) if tax.amount else 0
                tax_key = f"{tax_rate}%"

                # Calcular base gravable y monto de impuesto para esta línea
                line_base = abs(line.price_subtotal) * multiplier
                line_tax = abs(line.price_subtotal * (tax.amount / 100)) * multiplier

                if tax_key not in tax_breakdown:
                    tax_breakdown[tax_key] = {'base': 0, 'tax': 0}

                tax_breakdown[tax_key]['base'] += line_base
                tax_breakdown[tax_key]['tax'] += line_tax

        return tax_breakdown

    def _get_base_domain(self):
        """Returns base domain for purchase invoices"""
        domain = [
            ('invoice_date', '>=', self.start_date),
            ('invoice_date', '<=', self.end_date),
            ('company_id.id', 'in', self.company_ids.ids),
            ('state', '=', 'posted'),
            ('move_type', 'in', ['in_invoice', 'in_refund']),
        ]
        return domain

    def get_report_datas_by_invoice(self):
        """
        Obtiene datos por factura individual con columnas separadas por tasa de impuesto
        Muestra nombre de proveedor y número de factura en cada línea
        """
        domain = self._get_base_domain()
        if self.partner_ids:
            domain.append(('partner_id.id', 'in', self.partner_ids.ids))

        invoice_lines = self.env['account.move'].sudo().search(domain, order='partner_id, invoice_date')
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("There are not any invoices, Please review your selection."))

        # Lista para almacenar datos de cada factura
        invoice_data_list = []

        for inv in invoice_lines:
            # Obtener desglose de impuestos para esta factura
            tax_breakdown = self._get_invoice_tax_breakdown(inv)

            # Calcular subtotal (suma de todas las bases gravables)
            subtotal = sum(tax_data['base'] for tax_data in tax_breakdown.values())
            total_tax = sum(tax_data['tax'] for tax_data in tax_breakdown.values())
            total = subtotal + total_tax

            # Datos de la factura
            invoice_data = {
                'partner_name': inv.partner_id.name,
                'ref': inv.ref or inv.name,
                'date': inv.invoice_date,
                'currency': inv.currency_id.name,
                'is_usd': inv.currency_id.name == 'USD',
                'subtotal': subtotal,
                'total': total,
                'tax_breakdown': tax_breakdown,
            }

            invoice_data_list.append(invoice_data)

        return invoice_data_list

    def get_report_datas_partners(self):
        domain = self._get_base_domain()
        if self.partner_ids:
            domain.append(('partner_id.id', 'in', self.partner_ids.ids))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("There are not any invoices, Please review your selection."))

        partner_list = invoice_lines.mapped('partner_id').ids
        partners = self.env['res.partner'].sudo().search([('id', 'in', partner_list)])

        data_dics = []
        for p in partners:
            invoice = invoice_lines.filtered(lambda a: a.partner_id.id == p.id)
            bal_total_untax = 0
            total_tax_paid = 0
            balance_total_tax = 0
            tax_groups = []
            partner_dic = {}
            tex_gr_names = []  # Resetear para cada partner para evitar duplicados

            for inv in invoice:
                in_out_refund = inv.move_type in ['in_refund']
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        # Valores positivos para compras normales, negativos para notas de crédito
                        bal_total_untax += abs(inv.amount_untaxed) if not in_out_refund else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            tax_amount = abs(tax_group.get('tax_amount')) if not in_out_refund else -abs(tax_group.get('tax_amount'))
                            balance_total_tax += tax_amount
                            total_tax_paid += tax_amount

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name, invoice),
                })

            partner_dic['name'] = p.name if p.name else 'Name Not Available'
            partner_dic['bal_total_untax'] = bal_total_untax
            partner_dic['total_tax_paid'] = total_tax_paid
            partner_dic['balance_total_tax'] = balance_total_tax
            partner_dic['tax_groups'] = tax_groups
            data_dics.append(partner_dic)

        return data_dics

    def get_report_datas_taxes(self):
        domain = self._get_base_domain()
        if self.tax_ids:
            domain.append(('invoice_line_ids.tax_ids', 'in', self.tax_ids.ids))

        invoice_lines = self.env['account.move'].sudo().search(domain)
        if len(invoice_lines) == 0:
            raise models.ValidationError(_("There are not any invoices, Please review your selection."))

        tax_lst = invoice_lines.mapped('invoice_line_ids').mapped('tax_ids').ids
        tax_ids = self.env['account.tax'].sudo().search([('id', 'in', tax_lst)])

        data_dics = []

        for tax in tax_ids:
            invoice = invoice_lines.filtered(lambda a: tax in a.invoice_line_ids.tax_ids)
            bal_total_untax = 0
            total_tax_paid = 0
            balance_total_tax = 0
            tax_groups = []
            tax_dic = {}
            tex_gr_names = []  # Resetear para cada tax para evitar duplicados

            for inv in invoice:
                in_out_refund = inv.move_type in ['in_refund']
                taxes = inv.tax_totals
                if taxes.get('subtotals'):
                    if taxes['subtotals'][0]['tax_groups']:
                        tax_groups_lst = taxes['subtotals'][0]['tax_groups']
                        # Valores positivos para compras normales, negativos para notas de crédito
                        bal_total_untax += abs(inv.amount_untaxed) if not in_out_refund else -abs(inv.amount_untaxed)
                        for tax_group in tax_groups_lst:
                            tex_gr_names.append(tax_group.get('group_name'))
                            tax_amount = abs(tax_group.get('tax_amount')) if not in_out_refund else -abs(tax_group.get('tax_amount'))
                            balance_total_tax += tax_amount
                            total_tax_paid += tax_amount

            tex_gr_names_set = list(set(tex_gr_names))
            for tex_gr_name in tex_gr_names_set:
                tax_groups.append({
                    'tax': tex_gr_name,
                    'lines': self._filter_inv_on_tax(tex_gr_name, invoice),
                })

            tax_dic['name'] = tax.name if tax.name else 'Name Not Available'
            tax_dic['bal_total_untax'] = bal_total_untax
            tax_dic['total_tax_paid'] = total_tax_paid
            tax_dic['balance_total_tax'] = balance_total_tax
            tax_dic['tax_groups'] = tax_groups
            data_dics.append(tax_dic)

        return data_dics

    def _write_cell(self, worksheet, row, col, value, style_name):
        """Helper to write cell with style"""
        cell = worksheet.cell(row=row, column=col)
        cell.value = value

        if 'header' in style_name:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        elif 'bold' in style_name:
            cell.font = Font(bold=True)

        # Aplicar color verde para facturas en USD
        if 'green' in style_name:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        if 'center' in style_name:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif 'right' in style_name:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif 'left' in style_name:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    def print_xlsx_report(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Purchases Tax Report'

        row = 1
        col = 1

        # Title
        self._write_cell(worksheet, row, col, 'Reporte de Impuestos de Compras', 'header_center_bold')
        row += 2

        # Date range
        self._write_cell(worksheet, row, col, f'Desde: {self.start_date} Hasta: {self.end_date}', 'left_bold')
        row += 2

        if self.report_for == 'taxes':
            # Obtener datos por factura individual
            invoice_data_list = self.get_report_datas_by_invoice()

            # Determinar todas las tasas de impuesto únicas
            all_tax_rates = set()
            for invoice_data in invoice_data_list:
                all_tax_rates.update(invoice_data['tax_breakdown'].keys())

            all_tax_rates = sorted(all_tax_rates, key=lambda x: int(x.replace('%', '')))

            # Escribir headers
            current_col = col
            self._write_cell(worksheet, row, current_col, 'Fecha', 'header_center_bold')
            current_col += 1
            self._write_cell(worksheet, row, current_col, 'Nombre Completo', 'header_center_bold')
            current_col += 1
            self._write_cell(worksheet, row, current_col, 'Número Factura', 'header_center_bold')
            current_col += 1
            self._write_cell(worksheet, row, current_col, 'Moneda', 'header_center_bold')
            current_col += 1

            # Crear headers para cada tasa de impuesto (Base y Tarifa)
            for tax_rate in all_tax_rates:
                self._write_cell(worksheet, row, current_col, f'Grav {tax_rate}', 'header_center_bold')
                current_col += 1
                self._write_cell(worksheet, row, current_col, f'Tarifa {tax_rate}', 'header_center_bold')
                current_col += 1

            self._write_cell(worksheet, row, current_col, 'Sub Total', 'header_center_bold')
            current_col += 1
            self._write_cell(worksheet, row, current_col, 'Total Comprobant', 'header_center_bold')
            row += 1

            # Totales generales
            grand_totals = {'sub_total': 0, 'total': 0}
            for tax_rate in all_tax_rates:
                grand_totals[f'{tax_rate}_base'] = 0
                grand_totals[f'{tax_rate}_tax'] = 0

            # Escribir datos por factura
            for invoice_data in invoice_data_list:
                current_col = col

                # Determinar el estilo base según si es USD
                base_style = 'green_' if invoice_data['is_usd'] else ''

                # Fecha de factura
                self._write_cell(worksheet, row, current_col, str(invoice_data['date']), base_style + 'center')
                current_col += 1

                # Nombre del proveedor
                self._write_cell(worksheet, row, current_col, invoice_data['partner_name'], base_style + 'left')
                current_col += 1

                # Número de factura (ref)
                self._write_cell(worksheet, row, current_col, invoice_data['ref'], base_style + 'left')
                current_col += 1

                # Moneda
                self._write_cell(worksheet, row, current_col, invoice_data['currency'], base_style + 'center')
                current_col += 1

                # Escribir valores para cada tasa de impuesto
                for tax_rate in all_tax_rates:
                    if tax_rate in invoice_data['tax_breakdown']:
                        base = invoice_data['tax_breakdown'][tax_rate]['base']
                        tax = invoice_data['tax_breakdown'][tax_rate]['tax']
                        self._write_cell(worksheet, row, current_col, base, base_style + 'right')
                        current_col += 1
                        self._write_cell(worksheet, row, current_col, tax, base_style + 'right')
                        current_col += 1

                        # Acumular totales
                        grand_totals[f'{tax_rate}_base'] += base
                        grand_totals[f'{tax_rate}_tax'] += tax
                    else:
                        # Si no hay impuesto de esta tasa, dejar en blanco
                        self._write_cell(worksheet, row, current_col, '-', base_style + 'center')
                        current_col += 1
                        self._write_cell(worksheet, row, current_col, '-', base_style + 'center')
                        current_col += 1

                # Escribir subtotal y total
                self._write_cell(worksheet, row, current_col, invoice_data['subtotal'], base_style + 'right')
                current_col += 1
                self._write_cell(worksheet, row, current_col, invoice_data['total'], base_style + 'right')
                row += 1

                # Acumular totales generales
                grand_totals['sub_total'] += invoice_data['subtotal']
                grand_totals['total'] += invoice_data['total']

            # Escribir fila de TOTALES GENERALES
            row += 1
            current_col = col
            self._write_cell(worksheet, row, current_col, '', 'header_center_bold')  # Columna de fecha vacía
            current_col += 1
            self._write_cell(worksheet, row, current_col, 'TOTAL GENERAL', 'header_center_bold')
            current_col += 1
            self._write_cell(worksheet, row, current_col, '', 'header_center_bold')  # Columna de factura vacía
            current_col += 1
            self._write_cell(worksheet, row, current_col, '', 'header_center_bold')  # Columna de moneda vacía
            current_col += 1

            for tax_rate in all_tax_rates:
                self._write_cell(worksheet, row, current_col, grand_totals[f'{tax_rate}_base'], 'header_right')
                current_col += 1
                self._write_cell(worksheet, row, current_col, grand_totals[f'{tax_rate}_tax'], 'header_right')
                current_col += 1

            self._write_cell(worksheet, row, current_col, grand_totals['sub_total'], 'header_right')
            current_col += 1
            self._write_cell(worksheet, row, current_col, grand_totals['total'], 'header_right')

        elif False:  # Deshabilitar código antiguo
            for rec in data:
                self._write_cell(worksheet, row, col, rec.get('name'), 'left_bold')
                self._write_cell(worksheet, row, col+1, rec.get('bal_total_untax'), 'right_bold') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+2, rec.get('total_tax_paid'), 'right_bold') if not self.detailed_report else None
                self._write_cell(worksheet, row, col+3, rec.get('balance_total_tax'), 'right_bold') if not self.detailed_report else None
                row += 1

                for tex_line in rec.get('tax_groups'):
                    if self.detailed_report:
                        self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left_bold')
                        row += 1

                        # Headers específicos para COMPRAS (SIN IVA Devuelto)
                        self._write_cell(worksheet, row, col+2, _("Proveedor"), 'header_center_bold')
                        self._write_cell(worksheet, row, col+3, _("Factura"), 'header_center_bold')
                        self._write_cell(worksheet, row, col+4, _("Fecha Factura"), 'header_center_bold')
                        self._write_cell(worksheet, row, col+5, _("Sub Total"), 'header_right')
                        self._write_cell(worksheet, row, col+6, _("Descuento Factura Electrónica"), 'header_right')
                        self._write_cell(worksheet, row, col+7, _("Base Gravable"), 'header_right')
                        self._write_cell(worksheet, row, col+8, _("IVA Soportado"), 'header_right')
                        self._write_cell(worksheet, row, col+9, _("Total Firmado"), 'header_right')
                        self._write_cell(worksheet, row, col+10, _("Estado Tributación"), 'header_center_bold')
                        row += 1

                        # Initialize totals
                        totals = {
                            'sub_total': 0,
                            'amount_discount_electronic_invoice': 0,
                            'amount_untaxed': 0,
                            'amount_tax_signed': 0,
                            'amount_total_signed': 0,
                        }

                        for line in tex_line.get('lines'):
                            self._write_cell(worksheet, row, col+2, line.get('name'), 'left')
                            self._write_cell(worksheet, row, col+3, line.get('ref'), 'left')
                            self._write_cell(worksheet, row, col+4, line.get('inv_date'), 'left')
                            self._write_cell(worksheet, row, col+5, line.get('sub_total', 0), 'right')
                            self._write_cell(worksheet, row, col+6, line.get('amount_discount_electronic_invoice', 0), 'right')
                            self._write_cell(worksheet, row, col+7, line.get('amount_untaxed', 0), 'right')
                            self._write_cell(worksheet, row, col+8, line.get('amount_tax_signed', 0), 'right')
                            self._write_cell(worksheet, row, col+9, line.get('amount_total_signed', 0), 'right')
                            self._write_cell(worksheet, row, col+10, line.get('state_tributacion', ''), 'left')

                            # Update totals del grupo
                            for key in totals.keys():
                                totals[key] += line.get(key, 0)

                            # Update totales generales
                            for key in grand_totals.keys():
                                grand_totals[key] += line.get(key, 0)

                            row += 1

                        # Write totals row
                        self._write_cell(worksheet, row, col+4, 'Total:', 'right_bold')
                        self._write_cell(worksheet, row, col+5, totals['sub_total'], 'right_bold')
                        self._write_cell(worksheet, row, col+6, totals['amount_discount_electronic_invoice'], 'right_bold')
                        self._write_cell(worksheet, row, col+7, totals['amount_untaxed'], 'right_bold')
                        self._write_cell(worksheet, row, col+8, totals['amount_tax_signed'], 'right_bold')
                        self._write_cell(worksheet, row, col+9, totals['amount_total_signed'], 'right_bold')
                        row += 1
                    else:
                        self._write_cell(worksheet, row, col+1, tex_line.get('tax'), 'left')
                        total_base = 0
                        total_tax = 0
                        for line in tex_line.get('lines'):
                            total_base += line.get('base_amount', 0) or 0
                            total_tax += line.get('tax_amount', 0) or 0
                        self._write_cell(worksheet, row, col+2, total_base, 'right')
                        self._write_cell(worksheet, row, col+3, total_tax, 'right')
                        row += 1
                row += 1

            # Escribir fila de TOTALES GENERALES al final (solo en reporte detallado)
            if self.detailed_report:
                row += 1
                # Aplicar color de fondo amarillo para resaltar la fila de totales
                self._write_cell(worksheet, row, col+2, 'TOTAL GENERAL', 'header_center_bold')
                self._write_cell(worksheet, row, col+5, grand_totals['sub_total'], 'header_right')
                self._write_cell(worksheet, row, col+6, grand_totals['amount_discount_electronic_invoice'], 'header_right')
                self._write_cell(worksheet, row, col+7, grand_totals['amount_untaxed'], 'header_right')
                self._write_cell(worksheet, row, col+8, grand_totals['amount_tax_signed'], 'header_right')
                self._write_cell(worksheet, row, col+9, grand_totals['amount_total_signed'], 'header_right')

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)

        export_id = self.env['account.excel.tax.report'].create({
            'excel_file': base64.b64encode(fp.getvalue()),
            'file_name': 'purchases_tax_report.xlsx'
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'account.excel.tax.report',
            'view_mode': 'form',
            'res_id': export_id.id,
            'views': [(False, 'form')],
            'target': 'new',
        }

    def print_report(self):
        self.ensure_one()
        if self.report_format == 'excel':
            return self.print_xlsx_report()
        elif self.report_format == 'pdf':
            # TODO: Implement PDF report
            raise models.UserError(_("PDF report not implemented yet"))
