# -*- coding: utf-8 -*-

import base64
import csv
import io
import re
import unicodedata
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher

import openpyxl

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools import float_is_zero

HEADER_ALIASES = {
    "date": {"fecha", "fecha transaccion", "fecha de transaccion", "transaction date", "date"},
    "amount": {"monto", "importe", "valor", "amount", "abono", "deposito"},
    "credit": {"creditos", "credito", "haber"},
    "debit": {"debitos", "debito", "cargos", "retiros", "debe"},
    "reference": {"referencia", "reference", "numero de referencia", "ref", "numero de documento"},
    "description": {"descripcion", "detalle", "concepto", "memo", "description"},
    "partner_name": {"beneficiario", "cliente", "nombre", "ordenante", "partner", "payer"},
}
# Cuantas filas iniciales se revisan buscando la fila de encabezados reales de la
# tabla de movimientos, ya que muchos bancos (p.ej. BAC) anteponen un bloque de
# metadatos (nombre de cuenta, saldos, etc.) antes de la tabla.
HEADER_SEARCH_LIMIT = 50


class SngBankStatementMatchWizard(models.TransientModel):
    _name = "sng.bank.statement.match.wizard"
    _description = "Conciliacion masiva de pagos borrador vs extracto bancario"

    company_id = fields.Many2one(
        "res.company", string="Compania",
        help="Opcional. Deja vacio para comparar contra pagos borrador de todas las "
             "companias permitidas en tu sesion (recomendado si el banco cubre varias companias).",
    )
    file = fields.Binary(
        string="Extracto bancario",
        help="Sube un archivo .xlsx o .csv con columnas de Fecha, Monto y Referencia.",
    )
    filename = fields.Char(string="Nombre del archivo")
    journal_id = fields.Many2one(
        "account.journal", string="Diario bancario",
        domain="[('type', 'in', ('bank', 'cash'))]",
        help="Opcional. Si se indica, solo se consideran pagos borrador de este diario.",
    )
    date_tolerance_days = fields.Integer(string="Tolerancia de dias", default=3)
    state = fields.Selection(
        [("upload", "Cargar archivo"), ("review", "Revisar coincidencias")],
        default="upload",
    )
    line_ids = fields.One2many("sng.bank.statement.match.line", "wizard_id", string="Lineas del extracto")
    matched_count = fields.Integer(compute="_compute_counts")
    ambiguous_count = fields.Integer(compute="_compute_counts")
    no_match_count = fields.Integer(compute="_compute_counts")

    @api.depends("line_ids.match_status")
    def _compute_counts(self):
        for wizard in self:
            wizard.matched_count = len(wizard.line_ids.filtered(lambda l: l.match_status == "matched"))
            wizard.ambiguous_count = len(wizard.line_ids.filtered(lambda l: l.match_status == "ambiguous"))
            wizard.no_match_count = len(wizard.line_ids.filtered(lambda l: l.match_status == "no_match"))

    # ---------------------------------------------------------
    # Parsing helpers
    # ---------------------------------------------------------

    @staticmethod
    def _normalize_header(value):
        text = unicodedata.normalize("NFKD", str(value or "").strip())
        text = "".join(char for char in text if not unicodedata.combining(char))
        return re.sub(r"\s+", " ", text).lower()

    @staticmethod
    def _parse_number(value):
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(" ", "").replace("₡", "").replace("$", "")
        if not text:
            return None
        if "," in text and "." in text:
            if re.match(r"^-?\d{1,3}(\.\d{3})+,\d+$", text):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".") if text.count(",") == 1 else text.replace(",", "")
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _cell(row, index):
        return row[index] if index is not None and index < len(row) else None

    def _match_header(self, headers, key):
        for alias in HEADER_ALIASES[key]:
            if alias in headers:
                return headers[alias]
        return None

    def _find_header_row(self, raw_rows):
        """Busca, entre las primeras filas, la fila de encabezados reales de la
        tabla de movimientos (requiere Fecha + al menos una columna de monto).
        Muchos extractos bancarios (p.ej. BAC) anteponen un bloque de metadatos
        (nombre de cuenta, saldos) antes de la tabla real."""
        for row_index, row in enumerate(raw_rows[:HEADER_SEARCH_LIMIT]):
            headers = {
                self._normalize_header(value): index
                for index, value in enumerate(row)
                if value not in (None, "")
            }
            has_date = self._match_header(headers, "date") is not None
            has_amount = any(
                self._match_header(headers, key) is not None
                for key in ("amount", "credit", "debit")
            )
            if has_date and has_amount:
                return row_index, headers
        return None, None

    def _read_xlsx_rows(self):
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file)), data_only=True)
        except Exception as error:
            raise UserError(_("No se pudo leer el archivo Excel: %s") % error) from error
        sheet = workbook.active
        raw_rows = [
            row for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)
        ]
        return self._extract_headers_and_rows(raw_rows)

    def _read_csv_rows(self):
        raw_bytes = base64.b64decode(self.file)
        try:
            raw = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raw = raw_bytes.decode("latin-1")
        sample = raw[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        raw_rows = list(csv.reader(io.StringIO(raw), dialect))
        return self._extract_headers_and_rows(raw_rows)

    def _extract_headers_and_rows(self, raw_rows):
        if not raw_rows:
            raise UserError(_("El archivo esta vacio."))
        header_row_index, headers = self._find_header_row(raw_rows)
        if headers is None:
            raise UserError(
                _("No se encontro una fila de encabezados con Fecha y Monto (o Debitos/Creditos) "
                  "en las primeras %s filas del archivo.") % HEADER_SEARCH_LIMIT
            )
        data_rows = [
            row for row in raw_rows[header_row_index + 1:]
            if row and any(value not in (None, "") for value in row)
        ]
        return headers, data_rows

    def _load_rows(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Sube un archivo para continuar."))
        filename = (self.filename or "").lower()
        if filename.endswith(".csv"):
            headers, rows = self._read_csv_rows()
        elif filename.endswith(".xlsx"):
            headers, rows = self._read_xlsx_rows()
        else:
            raise UserError(_("El archivo debe estar en formato .xlsx o .csv."))

        date_index = self._match_header(headers, "date")
        amount_index = self._match_header(headers, "amount")
        credit_index = self._match_header(headers, "credit")
        debit_index = self._match_header(headers, "debit")
        reference_index = self._match_header(headers, "reference")
        description_index = self._match_header(headers, "description")
        partner_index = self._match_header(headers, "partner_name")
        dual_column_mode = credit_index is not None or debit_index is not None

        if date_index is None or (amount_index is None and not dual_column_mode):
            detected = " | ".join(headers.keys()) or _("sin encabezados")
            raise UserError(
                _("No se encontraron las columnas de Fecha y Monto (o Debitos/Creditos). "
                  "Encabezados detectados: %s") % detected
            )

        errors = []
        parsed_rows = []
        for row_number, row in enumerate(rows, start=1):
            row_date = self._parse_date(self._cell(row, date_index))
            if row_date is None:
                # Fila que no es un movimiento (saldo inicial, resumen, totales, etc.)
                continue

            if dual_column_mode:
                credit_val = self._parse_number(self._cell(row, credit_index)) if credit_index is not None else None
                debit_val = self._parse_number(self._cell(row, debit_index)) if debit_index is not None else None
                row_amount = credit_val or debit_val
                if not row_amount:
                    continue
            else:
                row_amount = self._parse_number(self._cell(row, amount_index))
                if row_amount is None:
                    errors.append(_("Fila %s: el monto no es valido.") % row_number)
                    continue

            reference_text = " ".join(
                str(self._cell(row, idx)).strip()
                for idx in (reference_index, description_index)
                if idx is not None and self._cell(row, idx) not in (None, "")
            )
            parsed_rows.append({
                "date": row_date,
                "amount": abs(row_amount),
                "reference": reference_text,
                "partner_name": str(self._cell(row, partner_index) or "").strip(),
            })

        if errors:
            raise UserError(_("El archivo tiene errores:\n%s") % "\n".join(errors[:30]))
        if not parsed_rows:
            raise UserError(_("El archivo no contiene lineas validas para importar."))
        return parsed_rows

    # ---------------------------------------------------------
    # Matching
    # ---------------------------------------------------------

    def _find_matching_payment(self, row, draft_payments):
        tolerance = timedelta(days=self.date_tolerance_days or 0)
        candidates = draft_payments.filtered(
            lambda p: float_is_zero(p.amount - row["amount"], precision_digits=2)
            and (row["date"] - tolerance) <= p.date <= (row["date"] + tolerance)
        )
        if len(candidates) == 1:
            return candidates, "matched"
        if len(candidates) > 1:
            needle = self._normalize_header("%s %s" % (row["reference"], row["partner_name"]))
            scored = []
            for payment in candidates:
                haystack = self._normalize_header(
                    "%s %s %s" % (payment.partner_id.name or "", payment.memo or "", payment.payment_reference or "")
                )
                score = SequenceMatcher(None, needle, haystack).ratio() if needle and haystack else 0.0
                scored.append((score, payment))
            scored.sort(key=lambda item: item[0], reverse=True)
            best_score, best_payment = scored[0]
            second_score = scored[1][0] if len(scored) > 1 else 0.0
            if best_score >= 0.6 and (best_score - second_score) >= 0.15:
                return best_payment, "matched"
            return candidates, "ambiguous"
        return self.env["account.payment"], "no_match"

    def action_import(self):
        self.ensure_one()
        rows = self._load_rows()

        domain = [("state", "=", "draft")]
        if self.journal_id:
            domain.append(("journal_id", "=", self.journal_id.id))

        if self.company_id:
            domain.append(("company_id", "=", self.company_id.id))
            payment_model = self.env["account.payment"]
        else:
            # El selector de companias del navegador (allowed_company_ids) puede tener
            # activa una sola compania aunque el usuario tenga permiso sobre varias.
            # Buscamos en todas las companias que el usuario tiene permitidas, no solo
            # la(s) activa(s) en ese momento, para no perder coincidencias validas.
            payment_model = self.env["account.payment"].with_context(
                allowed_company_ids=self.env.user.company_ids.ids
            )
        draft_payments = payment_model.search(domain)

        line_vals = []
        used_payment_ids = set()
        for row in rows:
            available = draft_payments.filtered(lambda p: p.id not in used_payment_ids)
            result, status = self._find_matching_payment(row, available)
            payment_id = False
            if status == "matched":
                payment_id = result.id
                used_payment_ids.add(result.id)
            line_vals.append({
                "date": row["date"],
                "amount": row["amount"],
                "reference": row["reference"],
                "partner_name": row["partner_name"],
                "payment_id": payment_id,
                "match_status": status,
                "apply": status == "matched",
            })

        self.write({
            "line_ids": [(5, 0, 0)] + [(0, 0, vals) for vals in line_vals],
            "state": "review",
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": dict(self.env.context, dialog_size="extra-large"),
        }

    # ---------------------------------------------------------
    # Apply
    # ---------------------------------------------------------

    def action_apply(self):
        self.ensure_one()
        to_apply = self.line_ids.filtered(lambda l: l.apply and l.payment_id)

        if not to_apply:
            raise UserError(_("No hay lineas seleccionadas para aplicar."))

        seen_payments = set()
        for line in to_apply:
            if line.payment_id.id in seen_payments:
                raise UserError(
                    _("El pago %s esta asignado a mas de una linea del extracto. Corrige antes de continuar.")
                    % line.payment_id.name
                )
            seen_payments.add(line.payment_id.id)

        applied = 0
        failed = []
        for line in to_apply:
            payment = line.payment_id
            if payment.state != "draft":
                failed.append(_("%s: ya no esta en borrador.") % payment.name)
                continue
            try:
                with self.env.cr.savepoint():
                    payment.action_post()
                applied += 1
            except Exception as error:  # noqa: BLE001 - se reporta al usuario, no debe detener el lote
                failed.append(_("%s: %s") % (payment.name, error))

        message = _("%s pago(s) aplicado(s) correctamente.") % applied
        if failed:
            message += "\n" + _("No se pudieron aplicar:") + "\n" + "\n".join(failed)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Exito") if not failed else _("Aplicado con advertencias"),
                "message": message,
                "type": "success" if not failed else "warning",
                "sticky": bool(failed),
                "next": {"type": "ir.actions.act_window_close"},
            },
        }


class SngBankStatementMatchLine(models.TransientModel):
    _name = "sng.bank.statement.match.line"
    _description = "Linea de conciliacion masiva de pagos borrador"
    _order = "id"

    wizard_id = fields.Many2one("sng.bank.statement.match.wizard", ondelete="cascade", required=True)
    date = fields.Date(string="Fecha extracto")
    amount = fields.Float(string="Monto extracto", digits=(16, 2))
    reference = fields.Char(string="Referencia extracto")
    partner_name = fields.Char(string="Beneficiario extracto")
    payment_id = fields.Many2one(
        "account.payment", string="Pago borrador", domain="[('state', '=', 'draft')]"
    )
    payment_date = fields.Date(related="payment_id.date", string="Fecha pago", readonly=True)
    payment_amount = fields.Monetary(related="payment_id.amount", string="Monto pago", readonly=True)
    payment_partner_id = fields.Many2one(
        related="payment_id.partner_id", string="Cliente/Proveedor pago", readonly=True
    )
    currency_id = fields.Many2one(related="payment_id.currency_id", readonly=True)
    match_status = fields.Selection(
        [("matched", "Coincide"), ("ambiguous", "Ambigua"), ("no_match", "Sin coincidencia")],
        default="no_match", readonly=True,
    )
    apply = fields.Boolean(string="Aplicar")
